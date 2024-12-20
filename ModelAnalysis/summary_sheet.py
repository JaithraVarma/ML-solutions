import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def summary(config):
    file_path = os.path.join(config['out_dir'],"model_analysis.xlsx")
    model_info = pd.read_excel(file_path, sheet_name='model_info')
    memory_analysis_summary = pd.read_excel(file_path, sheet_name='memory_analysis_summary')
    unique_model_operator = pd.read_excel(file_path, sheet_name='unique_model_operator')
    model_percentage_supported_ops = pd.read_excel(file_path, sheet_name='model_percentage_supported_ops')

    model_info['Input Count'] = pd.to_numeric(model_info['Input Count'], errors='coerce')
    model_info['Output Count'] = pd.to_numeric(model_info['Output Count'], errors='coerce')
    memory_analysis_summary['Spilling(IFM+OFM > 1.5MB)'] = pd.to_numeric(memory_analysis_summary['Spilling(IFM+OFM > 1.5MB)'], errors='coerce')
    memory_analysis_summary['Wts Super Itr(wts > 1MB)'] = pd.to_numeric(memory_analysis_summary['Wts Super Itr(wts > 1MB)'], errors='coerce')
    memory_analysis_summary['FM > 1.5 & Wts > 1MB'] = pd.to_numeric(memory_analysis_summary['FM > 1.5 & Wts > 1MB'], errors='coerce')
    model_percentage_supported_ops['percentage_supported_ops'] = pd.to_numeric(model_percentage_supported_ops['percentage_supported_ops'], errors='coerce')

    single_io = len(model_info[(model_info['Input Count'] == 1) & (model_info['Output Count'] == 1)])
    two_to_five_io = len(model_info[((model_info['Input Count'] >= 2) & (model_info['Input Count'] < 5)) | ((model_info['Output Count'] >= 2) & (model_info['Output Count'] < 5))])
    more_than_five_io = len(model_info[(model_info['Input Count'] > 5) | (model_info['Output Count'] > 5)])

    spilling = len(memory_analysis_summary[memory_analysis_summary['Spilling(IFM+OFM > 1.5MB)'] > 0])
    wts_gt_1mb = len(memory_analysis_summary[memory_analysis_summary['Wts Super Itr(wts > 1MB)'] > 0])
    fm_gt_1_5_and_wts_gt_1mb = len(memory_analysis_summary[(memory_analysis_summary['FM > 1.5 & Wts > 1MB'] > 0)])

    models_gt_80 = len(model_percentage_supported_ops[model_percentage_supported_ops['percentage_supported_ops'] > 80])
    models_gt_60 = len(model_percentage_supported_ops[(model_percentage_supported_ops['percentage_supported_ops'] > 60) & (model_percentage_supported_ops['percentage_supported_ops'] <= 80)])
    models_lt_30 = len(model_percentage_supported_ops[model_percentage_supported_ops['percentage_supported_ops'] < 30])
    models_lt_10 = len(model_percentage_supported_ops[model_percentage_supported_ops['percentage_supported_ops'] < 10])

    summary_data = {
        'Category': ['Inputs/Outputs', '', '',
                    'Memory', '', '',
                    'Partitioner Offloading', '', '', ''],
        'Description': ['single input', '2 to 5 inputs', '>5 inputs',
                        'spilling', 'wts > 1MB', 'fm > 1.5mb & wts > 1mb',
                        '>80%', '>60%', '< 30%', '<10%'],
        'Count': [single_io, two_to_five_io, more_than_five_io,
                spilling, wts_gt_1mb, fm_gt_1_5_and_wts_gt_1mb,
                models_gt_80, models_gt_60, models_lt_30, models_lt_10]
    }

    summary_df = pd.DataFrame(summary_data)

    summary_file_path = 'summary_model_analysis.xlsx'
    with pd.ExcelWriter(summary_file_path, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        unsupported_operators_df = unique_model_operator[unique_model_operator['is_supported'] == False][['model', 'operator']]
        unsupported_operators_df.to_excel(writer, sheet_name='Unsupported Operators', index=False)

    wb = openpyxl.load_workbook(summary_file_path)
    ws = wb['Summary']

    ws.merge_cells('A2:A4')
    ws.merge_cells('A5:A7')
    ws.merge_cells('A8:A11')

    wb.save(summary_file_path)

    wb_orig = openpyxl.load_workbook(file_path)
    ws_orig = wb_orig['unique_model_operator']

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in range(2, ws_orig.max_row + 1):
        if ws_orig[f'D{row}'].value == False:
            ws_orig[f'B{row}'].fill = yellow_fill

    wb_orig.save(file_path)

    print(f"Operators highlighted in '{file_path}'.")
    print(f"Summary saved in 'model_analysis.xlsx'.")
