import os
import openpyxl
import argparse
import json
import pandas as pd
import memory_analysis
import concat_parent_analysis
import operator_analysis
import summary_mem_analysis
import model_info
import program_mem_analysis
import onnx_gmac_gflop
import csv
import extract_shift_attribute_info
import summary_sheet

def merge_excel_files(files, error_file, out_file, config):
    merged_workbook = openpyxl.Workbook()
    try:
        for file_path in files:
            if os.path.exists(file_path):
                file_name = os.path.basename(file_path).split(".")[0]
                wb = openpyxl.load_workbook(file_path, read_only=True)
                for sheet_name in wb.sheetnames:
                    if sheet_name == "Sheet1":
                        new_sheet = merged_workbook.create_sheet(title=file_name)
                    else:
                        new_sheet = merged_workbook.create_sheet(title=sheet_name)
                    sheet = wb[sheet_name]
                    for row in sheet.iter_rows(values_only=True):
                        new_sheet.append(row)
        if 'Sheet' in merged_workbook.sheetnames:
            merged_workbook.remove(merged_workbook['Sheet'])

        if os.path.exists(error_file):
            df = pd.read_csv(error_file)
            df_sheet = merged_workbook.create_sheet(title='Error_Log')
            for r in df.values.tolist():
                df_sheet.append(r)
                
        # Adding config details to a new sheet
        config_sheet = merged_workbook.create_sheet(title='config_details')
        for key, value in config.items():
            config_sheet.append([key, value])
                
        merged_file_path = os.path.join(config['out_dir'], out_file)
        merged_workbook.save(merged_file_path)
    
    finally:
        merged_workbook.close()

def remove_excel_files(excel_files):
    for file_name in excel_files:
        if os.path.exists(file_name):
            os.remove(file_name)

def check_dconv(opr_list):
    return 'ConvD' in opr_list

def check_gconv(opr_list):
    return 'ConvG' in opr_list

def gen_summary(config, output_file):
    path = os.path.join(config['out_dir'], output_file)

    mem_sum = pd.read_excel(path, sheet_name='memory_analysis_summary')
    mem_sum = mem_sum[['model_name', 'FM > 1.5 & Wts > 1MB']]

    opr_sum = pd.read_excel(path, 'unique_model_operator')
    opr_sum = opr_sum[['model', 'operator']]
    opr_sum = opr_sum.groupby('model')['operator'].agg(list).reset_index()
    opr_sum['Need DConv'] = opr_sum['operator'].apply(check_dconv)
    opr_sum['Need Gconv'] = opr_sum['operator'].apply(check_gconv)
    opr_sum = opr_sum.drop(columns='operator')
    opr_sum = opr_sum.rename(columns={'model': 'model_name'})

    pm_sum = pd.read_excel(path, 'program_memory_analysis_summary') 
    pm_sum = pm_sum[['Models', 'Total PM']]
    pm_sum['PM>16KB'] = pm_sum['Total PM']/1024 > 16
    pm_sum = pm_sum.drop(columns='Total PM')
    pm_sum = pm_sum.rename(columns={'Models': 'model_name'})

    df = pd.merge(mem_sum, opr_sum, on='model_name', how='inner')
    df = pd.merge(df, pm_sum, on='model_name', how='inner')
    with pd.ExcelWriter(path, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name='requirements', index=False)
    return

def analyze_operators(config):
    operator_analysis.analyze_operators(config['input_dir'], config['out_dir'], config['dtype'], config['model_names'], config['model_names_dir'], config['reuse'])

def analyze_program_memory(config):
    program_mem_analysis.mem_analyser(config['input_dir'], config['operators_yaml_dir'], config['out_dir'], config['model_names'], config['model_names_dir'], config['reuse'])

def analyze_memory(config):
    memory_analysis.analyze_memory(config['input_dir'], config['out_dir'], config['optimization_flag'], config['dtype'], config['model_names'], config['model_names_dir'], config['reuse'])
    try:
        combined_data = pd.concat([pd.read_excel(os.path.join(config['out_dir'], 'mem_analysis', file)) for file in os.listdir(os.path.join(config['out_dir'], 'mem_analysis')) if file.endswith('.xlsx')], ignore_index=True)
        combined_data = combined_data.round(5)
        combined_data.to_excel('combined_mem_analysis.xlsx', index=False)
        summary_mem_analysis.summarize_mem_analysis(os.path.join(config['out_dir'], 'mem_analysis'))
        concat_parent_analysis.filter_excel('combined_mem_analysis.xlsx', 'concat_parent_mem_analysis.xlsx')
    except:
        pass

def analyze_io_shape(config):
    model_info.main(config)

def gmac_gflop_info(config):
    onnx_gmac_gflop.main(config)

def extract_shift_model_attribute(config):
    input_dir=config['input_dir']
    filename='shift_and_attribute_info'
    dtype=config['dtype']
    if dtype=='int8':
        extract_shift_attribute_info.generate_info_excel(input_dir,filename,dtype)
    elif dtype=='bf16':
        extract_shift_attribute_info.generate_info_excel(input_dir,filename,dtype)

def analyze_all(config):
    analyze_operators(config)
    analyze_memory(config)
    analyze_io_shape(config)
    analyze_program_memory(config)
    gmac_gflop_info(config)
    # extract_shift_model_attribute(config)

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", help="Path to the config json file", required=True)
    args = parser.parse_args()

    with open(args.config, 'r') as json_file:
        config = json.load(json_file)
    if not os.path.exists(config["out_dir"]):
        os.makedirs(config["out_dir"])
    csv_file_path = os.path.join(config['out_dir'], 'error.csv')
    with open(csv_file_path, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(['Analysis', 'Model_name', 'Error'])

    analysis_functions = {
        "operator_analysis": analyze_operators,
        "program_memory_analysis": analyze_program_memory,
        "memory_analysis": analyze_memory,
        "model_info": analyze_io_shape,
        "shift_model_attribute": extract_shift_model_attribute,
        "onnx_gmacs_gflops": gmac_gflop_info,
        "": analyze_all,  # Default option to run all analyses
    }

    output_file = 'model_analysis.xlsx'
    analysis_type = config.get("analysis", "")

    if analysis_type in analysis_functions:
        analysis_functions[analysis_type](config)
    else:
        raise KeyError(f"Invalid analysis type '{analysis_type}'. Valid options are: operator_analysis, memory_analysis, model_info, program_memory_analysis, shift_model_attribute, onnx_gmacs_gflops, or an empty string for all analyses.")

    error_file = os.path.join(config["out_dir"], "error.csv")
    model_analysis_file_path = os.path.join(config["out_dir"], "model_analysis.xlsx")
    merge_excel_files(["combined_mem_analysis.xlsx", "concat_parent_mem_analysis.xlsx", "memory_analysis_summary.xlsx", "model_info.xlsx", "operator_analysis.xlsx", "program_mem_analysis.xlsx", "onnx_gmacs_gflops.xlsx"], error_file, output_file, config)
    summary_sheet.summary(config)
    merge_excel_files([model_analysis_file_path, "summary_model_analysis.xlsx"], error_file, output_file, config)
    gen_summary(config, output_file)
    remove_excel_files(["combined_mem_analysis.xlsx", "concat_parent_mem_analysis.xlsx", "memory_analysis_summary.xlsx", "model_info.xlsx", "operator_analysis.xlsx", "program_mem_analysis.xlsx","shift_and_attribute_info.xlsx","summary_model_analysis.xlsx", "onnx_gmacs_gflops.xlsx"])
